#import xml.etree.ElementTree as ET
import lxml.etree as ET
from zipfile import ZipFile
import socket
import struct
import sys
#import os
from pathlib import *
import shutil

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment

from copy import copy
import time

IED = "{http://www.iec.ch/61850/2003/SCL}IED"
PRIVATE = "{http://www.iec.ch/61850/2003/SCL}Private"
ACCESSPOINT = "{http://www.iec.ch/61850/2003/SCL}AccessPoint"
SERVER = "{http://www.iec.ch/61850/2003/SCL}Server"
LDEVICE = "{http://www.iec.ch/61850/2003/SCL}LDevice"
LN0 = "{http://www.iec.ch/61850/2003/SCL}LN0"
DATASET = "{http://www.iec.ch/61850/2003/SCL}DataSet"
HEADER = "{http://www.iec.ch/61850/2003/SCL}Header"

LN = "{http://www.iec.ch/61850/2003/SCL}LN"
DOI = "{http://www.iec.ch/61850/2003/SCL}DOI"
DAI = "{http://www.iec.ch/61850/2003/SCL}DAI"

COMMUNICATION = '{http://www.iec.ch/61850/2003/SCL}Communication'
SUBNETWORK = "{http://www.iec.ch/61850/2003/SCL}SubNetwork"
CONNECTEDAP = "{http://www.iec.ch/61850/2003/SCL}ConnectedAP"
ADDRESS = "{http://www.iec.ch/61850/2003/SCL}Address"
GSE = "{http://www.iec.ch/61850/2003/SCL}GSE"
MINTIME = "{http://www.iec.ch/61850/2003/SCL}MinTime"
MAXTIME = "{http://www.iec.ch/61850/2003/SCL}MaxTime"
P = "{http://www.iec.ch/61850/2003/SCL}P"

GSECONTROL = "{http://www.iec.ch/61850/2003/SCL}GSEControl"
FCDA = "{http://www.iec.ch/61850/2003/SCL}FCDA"

INPUTS = "{http://www.iec.ch/61850/2003/SCL}Inputs"
EXTREF = "{http://www.iec.ch/61850/2003/SCL}ExtRef"

CIPA = "{http://www.iec.ch/61850/2003/SCL}CIPA"

HORIZ = 4 # координаты начала заполнения EXCEL
VERT = 4 # координаты начала заполнения EXCEL
HORIZ_IED_NAME = 3
HORIZ_DATA_SET = 2
VERT_IED_NAME = 3
VERT_DATA_SET = 2
VERT_INDS = 1

class Report_cipa():
    def __init__(self):
        self.err = set()
        self.ok_term = set()
    def make_report(self, path_report: Path):
        #print(len(err))
        self.err.discard(None)
        with open(path_report/"report_cipa.txt", 'w', encoding='utf-8') as file:
            file.write("Ненайденных терминалов: " + str(len(self.err)) + "\n")
            for i in self.err:
                file.write("\t" + i + "\n")
            file.write("Обработанных терминалов :" + str(len(self.ok_term)) + "\n")
            for i in self.ok_term:
                file.write("\t" + i + "\n")

def whose_cid(cid: str):
    # print("whose_cid")
    # #print("\n!!!", cid)
    # s_t_whose_cid = time.time()
    # print("start", "whose_cid")

    try:
        tree = ET.parse(cid)
    except:
        p = cid.rsplit("/", 2)
#        print("Не удалось открыть", cid)
        err = open (p[0] + '/' + "no_read.txt", 'a')
        err.write(p[-1])
        err.write("\n")
        err.close()
        return "Не удалось открыть"
    #print("Определяем чей cid ",cid)
    root = tree.getroot()
    #print(root.nsmap)
    try:
        manufacturer = root.find(f"{IED}").attrib["manufacturer"] == "EKRA"
    except:
        e_t_whose_cid = time.time()
        # print("end", "whose_cid", f"time={e_t_whose_cid - s_t_whose_cid}")

        #print("Неизвестное устройство")
        return "Неизвестное устройство"
    if manufacturer:
        e_t_whose_cid = time.time()
        # print("end", "whose_cid", f"time={e_t_whose_cid - s_t_whose_cid}")

        #print("This is EKRA")
        return "IEC61850_TERMINAL"
        return "EKRA"

    else:
        e_t_whose_cid = time.time()
        # print("end", "whose_cid", f"time={e_t_whose_cid - s_t_whose_cid}")

        #print("IEC61850_TERMINAL")
        return "IEC61850_TERMINAL"

def make_xl(cids, path):
    s_t_make_xl = time.time()
    print("start", "make_xl")

    wb = openpyxl.Workbook()
    ws = wb.active
    horiz = HORIZ + 1 # первый столбец для заполнения горизонтали терминалов
    vert = VERT + 1 # первая строка для заполнения вертикали терминалов
    STYLE = Alignment(horizontal=None, vertical="center")
    for t in cids:
        # print("t =", t)
        terminal = ET.parse(path.parent/f"{t}.cid").getroot()
        t_ied = t
        t_ip = terminal.find(f".//{COMMUNICATION}//{CONNECTEDAP}[@iedName='{t}']/{ADDRESS}/{P}[@type='IP']").text
        #print("заполняем  xl", t.goose_out_param[5])
        ws.cell(VERT, horiz).value = t_ied
        ws.cell(VERT, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
        #ws.cell(VERT - 2, horiz).value = t.communication["IP"] 
        #ws.cell(VERT - 2, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
        letter = ws.cell(VERT, horiz).column_letter # получаем букву текущей ячейки чтобы изменить ее ширину
        ws.column_dimensions[letter].width = 3 # меняем ширину текущего столбца
        horiz += 1
        start_group = ws.cell(VERT, horiz).column_letter # получаем букву текущей для группировки ячеек
        end_group = None # если у терминала не будет входящих гусей, то конец гусей не определиться и колонки не будут группироваться
        for extref in terminal.findall(f".//{IED}[@name='{t}']//{INPUTS}/{EXTREF}"): # бежим по всем входящим нашего терминала
        #for ind, goose in terminal.findall()
            ied_name = extref.get("iedName") # если iedName нет, значит гусь не используется, пропускаем
            if ied_name == None:
                continue
            serviceType = extref.get("serviceType") # если serviceType "SMV" (SV-поток), пропускаем
            if serviceType == "SMV":
                continue

            ws.cell(VERT-VERT_INDS, horiz).value = extref.get("intAddr")
            ws.cell(VERT-VERT_INDS, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)

            ws.cell(VERT-VERT_IED_NAME, horiz).value = t_ied #f"{extref.attrib['iedName']}"
            ws.cell(VERT-VERT_IED_NAME, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
            prefix_extref = extref.get('prefix', "")
            ws.cell(VERT-VERT_DATA_SET, horiz).value = f"{extref.attrib['iedName']}/{extref.attrib['ldInst']}/{prefix_extref}{extref.attrib['lnClass']}{extref.attrib['lnInst']}/{extref.attrib['doName']}/{extref.attrib['daName']}"
            ws.cell(VERT-VERT_DATA_SET, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
            ws.cell(VERT, horiz).value = f"{extref.attrib['intAddr']}"
            ied = terminal.find(f".//{IED}")
            
            if ied.get("manufacturer") == "EKRA":
                intAddr = extref.attrib['intAddr'].split('.')
                for ln in terminal.findall(f".//{LDEVICE}/{LN}"):
                    if ln.get("prefix", "") + ln.get("lnClass", "") + ln.get("inst", "") == intAddr[0]:
                        try:
                            ws.cell(VERT, horiz).value = ln.find(f".//{DOI}[@name='{intAddr[1]}']/{DAI}").get("desc")
                        except:
                            print("\ошибка чтения", t)
            elif ied.get("manufacturer") == "SIEMENS":
                intAddr = extref.attrib['intAddr']
                ln = extref.getparent().getparent()
                for doi in ln.findall(f"./{DOI}"):
                    if doi.attrib["name"] in intAddr:
                        ws.cell(VERT, horiz).value = doi.attrib["desc"]
                        break
            elif ied.get("manufacturer") == "RELEMATIKA":
                ws.cell(VERT, horiz).value = extref.attrib["desc"]
            ws.cell(VERT, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
            letter = ws.cell(VERT, horiz).column_letter # получаем букву текущей ячейки чтобы изменить ее ширину
            ws.column_dimensions[letter].width = 3 # меняем ширину текущего столбца
            end_group = ws.cell(VERT, horiz).column_letter
            horiz += 1
        if end_group:
            ws.column_dimensions.group(start_group, end_group, hidden=True)
        #ip = terminal.find(f".//{COMMUNICATION}//{CONNECTEDAP}[@iedName='{t}']/{ADDRESS}/{P}[@type='IP']").text
        ws.cell(vert, HORIZ).value = t_ied
        #ws.cell(vert, HORIZ - 2).value = t.communication["IP"] 
        vert += 1
        start_group = vert
        end_group = None
        for gse in terminal.findall(f".//{COMMUNICATION}//{CONNECTEDAP}[@iedName='{t}']/{GSE}"): # прохожу по всем исходящим гусям
            cbName = gse.get("cbName") # получаю cbName чтобы потом найти нужный datSet
            ldInst = gse.get("ldInst") # получаю ldInst чтобы потом найти нужный datSet
            gse_address_p = "" # параметры гуся для добавления инфы в excell
            minTime = gse.find(f"./{MINTIME}").text
            maxTime = gse.find(f"./{MAXTIME}").text
            for p in gse.findall(f"./{ADDRESS}/{P}"):
                gse_address_p += f"\n{p.attrib['type']}={p.text}"
            # print("cbName =", cbName)
            datSet = terminal.find(f".//{IED}[@name='{t}']//{ACCESSPOINT}//{SERVER}//{LDEVICE}[@inst='{ldInst}']/{LN0}/{GSECONTROL}[@name='{cbName}']").get("datSet") # получил имя DataSet с набором данных гуся
            # print("datSet =", datSet)
            data_set_start = vert
            end_group = vert
            for fcda in terminal.findall(f".//{IED}[@name='{t}']/{ACCESSPOINT}/{SERVER}/{LDEVICE}[@inst='{ldInst}']/{LN0}/{DATASET}[@name='{datSet}']/{FCDA}"): # бегу по исходящим сигналам гуся
                # print(fcda.attrib)
                ws.cell(vert, HORIZ-HORIZ_IED_NAME).value = t_ied
                prefix_fcda = fcda.get("prefix", "")
                lnInst = fcda.get("lnInst", "")
                ws.cell(vert, HORIZ-1).value = f"{t}/{fcda.attrib['ldInst']}/{prefix_fcda}{fcda.attrib['lnClass']}{lnInst}/{fcda.attrib['doName']}/{fcda.attrib['daName']}"
                ws.cell(vert, HORIZ-HORIZ_DATA_SET).value = f"{t_ip} \nldInst={ldInst}, cbName={cbName} (datSet={datSet})" + gse_address_p + f"\nminTime = {minTime}, maxTime = {maxTime}"
                ws.cell(vert, HORIZ-HORIZ_DATA_SET).alignment = STYLE
                doName = fcda.attrib["doName"]
                cipa_name_signal = terminal.find(f".//{IED}[@name='{t}']//{LDEVICE}[@inst='{fcda.get('ldInst')}']/{LN}[@prefix='{fcda.get('prefix')}'][@lnClass='{fcda.get('lnClass')}'][@inst='{fcda.get('lnInst')}']/{DOI}[@name='{doName}']") # нахожу узел на который подписан [@doName='{doName}']
                #print(test1, test2, cipa_name_signal.attrib)
                #cipa = ET.Element(CIPA)
                #cipa.attrib["IED"] = ied_name
                #cipa.attrib["name"] = cipa_name_signal.get("name")
                try:
                    desc_in_DOI = cipa_name_signal.get("desc")
                except:
                    desc_in_DOI = "Нет описания в cid"
                if desc_in_DOI:
                    ws.cell(vert, HORIZ).value = desc_in_DOI
                else:
                    try:
                        desc_in_DAI = cipa_name_signal.find(DAI).get("desc")
                        ws.cell(vert, HORIZ).value = desc_in_DAI
                    except:
                        ws.cell(vert, HORIZ).value = "Не удалось получить название сигнала"
                
                
                
                # try:
                #     ws.cell(vert, HORIZ).value = fcda.find(f"./{CIPA}").attrib["desc"]
                # except:
                #     ws.cell(vert, HORIZ).value = 'Проблема в 343  ws.cell(vert, HORIZ).value = fcda.find(f"./CIPA").attrib["desc"]'
                # ws.cell(vert, HORIZ - 1).value = signal
                # ws.cell(vert, HORIZ - 2).value = t.signal_names.get(int(signal), "В ЭКРАвских сидах не найти, бери мануал(((")
                
                end_group = vert
                vert += 1
            #print(t)
            ws.merge_cells(range_string=None, start_row=data_set_start, start_column=HORIZ-HORIZ_DATA_SET, end_row=end_group, end_column=HORIZ-HORIZ_DATA_SET) # объеденяем ячейки
        # for ind, signal in t.goose_out_data.items():
        #     ws.cell(vert, HORIZ).value = "ind_" + str(ind)
        #     ws.cell(vert, HORIZ - 1).value = signal
        #     ws.cell(vert, HORIZ - 2).value = t.signal_names.get(int(signal), "В ЭКРАвских сидах не найти, бери мануал(((")
        #     end_group = vert
        #     vert += 1
        if end_group:
            ws.row_dimensions.group(start_group, end_group, hidden=True) 
    ws.column_dimensions.group('B', ws.cell(VERT, HORIZ - 1).column_letter, hidden=True)
    ws.row_dimensions.group(1, VERT - 1, hidden=True)
    ws.column_dimensions[ws.cell(VERT, HORIZ - HORIZ_IED_NAME).column_letter].width = 22
    ws.column_dimensions[ws.cell(VERT, HORIZ - 1).column_letter].width = 5
    ws.column_dimensions[ws.cell(VERT, HORIZ).column_letter].width = 20
    ws.row_dimensions[VERT].height = 100
    ws.row_dimensions[VERT - 2].height = 110
    ws.freeze_panes = ws.cell(VERT + 1, HORIZ + 1)
    wb.save(path.parent/"For_Sancho.xlsx")
    e_t_make_xl = time.time()
    print("end", "make_xl", f"time={e_t_make_xl - s_t_make_xl}")
    fill_xl(path.parent/"For_Sancho.xlsx")
    #to_vasiliy_xl(path.parent/"For_Sancho.xlsx", "vas.xlsx")

def fill_xl(file_xl):
    s_t_fill_xl = time.time()
    print("start", "fill_xl")
    wb = openpyxl.load_workbook(file_xl)
    ws = wb.active
    empty_vertical = [True] * (ws.max_row + 1)
    empty_horizontal = [True] * (ws.max_column + 1)
    max_column = ws.max_column
    horiz = HORIZ + 1 # первый столбец для заполнения пересечений
    vert = VERT + 1# первая строка для заполнения пересечений
    terminal_suppression = {} # ключи - названия терминалов записанные друг за друном, значения - адрес ячейки где терминалы пересекаются    
    BORDER_COLOR = 'C0C0C0' # Задаем цвет границ    
    CELL_COLOR = 'D3D3D3' # Задаем цвет границ    
    BORDER_SIDE = Side(style='thin', color=BORDER_COLOR)  # Создаем объект Side с нужным цветом    
    BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE) # Создаем объект Border с нужными границами
    print("...10...20...30...40...50...60...70...80...90..100%")
    PROTC = 2
    for x in range(horiz, ws.max_column + 1):
        if x / max_column * 100 >= PROTC:
            print("-", end="")
            sys.stdout.flush()
            PROTC += 2
        for y in range(vert, ws.max_row + 1):
            ws.cell(y, x).border = BORDER
            if ws.cell(y, HORIZ - 1).value == None:
                ws.cell(y, x).fill = PatternFill('solid', fgColor=CELL_COLOR)
            if ws.cell(VERT - 1, x).value == None:
                ws.cell(y, x).fill = PatternFill('solid', fgColor=CELL_COLOR)
                if ws.cell(y, HORIZ - 1).value == None:
                    # ws.cell(y, x).value = 'X'
                    terminal_suppression[ws.cell(VERT, x).value + ws.cell(y, HORIZ).value] = ws.cell(y, x)
                    continue
            #print("VERT =", ws.cell(y, HORIZ - 1).value, "HORIZ =", ws.cell(VERT - VERT_DATA_SET, x).value)    
            if ws.cell(y, HORIZ - 1).value == ws.cell(VERT - VERT_DATA_SET, x).value:
                ws.cell(y, x).value = 'x'
                row = terminal_suppression[ws.cell(VERT - VERT_IED_NAME, x).value + ws.cell(y, HORIZ - HORIZ_IED_NAME).value].row
                column = terminal_suppression[ws.cell(VERT - VERT_IED_NAME, x).value + ws.cell(y, HORIZ - HORIZ_IED_NAME).value].column
                #terminal_suppression[ws.cell(VERT - 2, x).value + ws.cell(y, HORIZ - 2).value].value = 'X'
                ws.cell(row, column).value = 'X'
                ws.cell(row, x).value = '->'
                ws.cell(row, x).alignment = openpyxl.styles.Alignment(textRotation=180)
                ws.cell(y, column).value = '->'

                ws.cell(row, HORIZ).fill = PatternFill('solid', fgColor=CELL_COLOR)
                ws.cell(VERT, column).fill = PatternFill('solid', fgColor=CELL_COLOR)
                
                continue
    # ws.delete_cols(1)
    # ws.columns.pop(1)
    wb.save(file_xl)
    e_t_fill_xl = time.time()
    print("\nend", "fill_xl", f"time={e_t_fill_xl - s_t_fill_xl}")
    
    #return empty_vertical, empty_horizontal

def paint_xl(file_xl, empty_vertical, empty_horizontal):    
    print("paint_xl")
    wb = openpyxl.load_workbook(file_xl)
    ws = wb.active
    horiz = HORIZ + 1 # первый столбец для заполнения пересечений
    vert = VERT + 1# первая строка для заполнения пересечений
    print("красим входящие, которые не подписаны")
    for x in range(horiz, ws.max_column + 1):        
        if empty_horizontal[x]:
            ws.cell(VERT, x).fill = PatternFill('solid', fgColor="8A3324")
    print("красим исходящие, на которые подписаны")
    for y in range(vert, ws.max_row + 1):        
        if empty_vertical[y]:
            ws.cell(y, HORIZ).fill = PatternFill('solid', fgColor="8A3324")
    wb.save(file_xl)

def to_vasiliy_xl (sourсe: Path, dest: Path):
    print("to_vasiliy_xl")
    # загружаем файлы
    source_wb = openpyxl.load_workbook(sourсe) # открываем файл-источник
    dest_wb = openpyxl.load_workbook(dest) # открываем файл назначения
    # выбираем листы для работы
    source_ws = source_wb.active # выбираем активный лист в файле-источнике
    dest_ws = dest_wb.worksheets[0] # выбираем первый лист в файле назначения
    # копируем значения и форматирование
    for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row,
                                    min_col=1, max_col=source_ws.max_column):  # проходимся по всем строкам в листе-источнике
        for cell in row:  # проходимся по всем ячейкам в текущей строке
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column)  # выбираем соответствующую ячейку в листе-назначении
            dest_cell.value = cell.value  # копируем значение ячейки
            dest_cell.data_type = cell.data_type  # копируем тип данных ячейки
            if cell.has_style:  # если ячейка имела стиль, копируем его в ячейку-назначение
                dest_cell.font = copy(cell.font)
                dest_cell.border = copy(cell.border)
                dest_cell.fill = copy(cell.fill)
                dest_cell.number_format = copy(cell.number_format)
                dest_cell.protection = copy(cell.protection)
                dest_cell.alignment = copy(cell.alignment)
    # сохраняем изменения в перезаписывая исходный файл
    dest_wb.save(sourсe)

def make_substation(cids: Path):
    s_t_make_substation = time.time()
    print("start", "make_substation")
    substation = {}
    report_cipa = Report_cipa()
    with ZipFile(cids, mode='r') as zip_file:
        info = zip_file.namelist() 
        zip_file.extractall(cids.parent)
        for file in info:
            manufacturer = "IEC61850_TERMINAL" #= whose_cid(cids.parent/file)
            if manufacturer == "IEC61850_TERMINAL":
                
                try:
                    tree = ET.parse(cids.parent/file)
                    root = tree.getroot()
                    ied_name = root.find(IED).attrib["name"] # получаю IED файла с которым работаю
                    # на этом месте нужно описась входящие и исходящие сигналы
                    substation[ied_name] = tree
                    print("OK  ", file)
                except:
                    print("ERR ",file)
            else:
                print("\n !!!!!!!!Пропускаем ", file)

    for ied_name, tree in substation.items():
        root = tree.getroot()
        # print(type(tree))
        # print(type(root))
        report_cipa.ok_term.add(ied_name) # отправили в отчет, что этот терминал обработан VV_T1
        for extref in root.findall(f".//{IED}[@name='{ied_name}']//{INPUTS}/{EXTREF}"): # бежим по всем входящим нашего терминала
            source_ied_name = extref.get("iedName") # выдергиваем ИЕД на который он подписан VV_T2
            if source_ied_name in substation: # если терминал с этим ИЕДом у нас есть, заходим в него и говорим что подписаны на него
                pass
            else:
                # print("НЕ нашли", source_ied_name)
                report_cipa.err.add(source_ied_name)
    
    for k, v in substation.items():
        v.write(cids.parent/f"{k}.cid", encoding="UTF-8")
        #zip_file.write(cids.parent/("c_" + k), (k + ".cid"))

    make_xl(substation.keys(), cids)

    with ZipFile(cids.parent/"cipa.zip", mode='w') as zip_file:
        # for k, v in substation.items():
        #     v.write(cids.parent/("c_" + k), encoding="UTF-8")
        #     zip_file.write(cids.parent/("c_" + k), (k + ".cid"))
        report_cipa.make_report(cids.parent)
        zip_file.write(cids.parent/("report_cipa.txt"), ("report_cipa.txt"))
        zip_file.write(cids.parent/"For_Sancho.xlsx", "For_Sancho.xlsx")
    e_t_make_substation = time.time()
    print("end", "make_substation", f"time={e_t_make_substation - s_t_make_substation}")


class SuperSocket():
    def __init__(self, sock):
        self._sock = sock
    def send_msg(self, msg):
        # Каждое сообщение будет иметь префикс в 4 байта блинной(network byte order)
        msg = struct.pack('>I', len(msg)) + msg
        self._sock.send(msg)
    def recv_msg(self):
        # Получение длины сообщения и распаковка в integer
        raw_msglen = self.recvall(4)
        if not raw_msglen:
            return None
        msglen = struct.unpack('>I', raw_msglen)[0]
        # Получение данных
        return self.recvall(msglen)
    def recvall(self, n):
        # Функция для получения n байт или возврата None если получен EOF
        data = b''
        while len(data) < n:
            packet = self._sock.recv(n - len(data))
            if not packet:
                return None
            data += packet
        return data

sock = socket.socket()
# IP = "213.178.155.215"
IP = ""
PORT = 51085
sock.bind((IP, PORT))
sock.listen(5)

while True:
    print("start")
    # начинаем принимать соединения
    conn, addr = sock.accept()
    # выводим информацию о подключении
    print('connected:', addr)
    super_sock = SuperSocket(conn)
     # получаем название файла
    name_f = super_sock.recv_msg().decode('UTF-8')
    # создаем название каталога клиента
    path_f = Path.cwd()/"clients"/name_f.removesuffix(".zip")
    # создаем каталог клиента если его нет
    if not Path.is_dir(path_f):
        Path.mkdir(path_f)
    # открываем файл в режиме байтовой записи в папке клиента и записываем его из сокета
    (path_f/name_f).write_bytes(super_sock.recv_msg())
    #f.close()

    print("Обрабытываем сиды")
    # make_substation("clients/" + name_f.strip(".zip") + "/" + name_f)
    make_substation(path_f/name_f)
    
    # name_f = "clients/" + name_f.strip(".zip") + "/" + name_f.strip(".zip") + ".txt"
    #name_f_client = path + name + ".xlsx"
    print("Отправляем", "test.xlsx")
    #super_sock.send_msg((f"{name_f.removesuffix('.zip')}.xml").encode('UTF-8'))
    #f = open (path_f/f"{name_f.removesuffix('.zip')}.xml", "rb")
    super_sock.send_msg(("cipa.zip").encode('UTF-8'))
    f = open (path_f/"cipa.zip", "rb")
    super_sock.send_msg(f.read())
    f.close()

    conn.close()
    print("end")
    print("убери break после отладки")
    break

