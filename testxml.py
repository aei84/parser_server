import xml.etree.ElementTree as ET
from zipfile import ZipFile
import socket
import struct
import sys
import os

import openpyxl
from openpyxl.styles import PatternFill

IED = "{http://www.iec.ch/61850/2003/SCL}IED"
PRIVATE = "{http://www.iec.ch/61850/2003/SCL}Private"
ACCESSPOINT = "{http://www.iec.ch/61850/2003/SCL}AccessPoint"
SERVER = "{http://www.iec.ch/61850/2003/SCL}Server"
LDEVICE = "{http://www.iec.ch/61850/2003/SCL}LDevice"
LN0 = "{http://www.iec.ch/61850/2003/SCL}LN0"
DATASET = "{http://www.iec.ch/61850/2003/SCL}DataSet"
HEADER = "{http://www.iec.ch/61850/2003/SCL}Header"

LN = "{http://www.iec.ch/61850/2003/SCL}LN"
DOI = "{http://www.iec.ch/61850/2003/SCL}DOI"
DAI = "{http://www.iec.ch/61850/2003/SCL}DAI"

COMMUNICATION = "{http://www.iec.ch/61850/2003/SCL}Communication"
SUBNETWORK = "{http://www.iec.ch/61850/2003/SCL}SubNetwork"
CONNECTEDAP = "{http://www.iec.ch/61850/2003/SCL}ConnectedAP"
ADDRESS = "{http://www.iec.ch/61850/2003/SCL}Address"
P = "{http://www.iec.ch/61850/2003/SCL}P"


class Terminal():
    def __init__(self):
        self.communication = {}
        self.goose_out_param = {} # {0: '1', 1: '010CCD010014', 2: '4', 3: '3', 4: '3114', 5: 'QC1G_F6_A2_PA', 6: '1', 7: '2.4', 8: '0'}
        self.goose_out_data = {} # {1: '443', 2: '444', 3: '363', 4: '356', 5: '357', 6: '360', 7: '329', 8: '223', 9: '154'}
        self.goose_in = {}
        self.subscribers = {} # {1: [('QC1G_A1_DD', 5), .... 7: [], 8: [], 9: []}
        self.errors = []
        self.signal_names = {} # {214: 'Готовность LAN1', 215: 'Готовность LAN2'...}
        self.ied = {}


def print_terminal(terminal: Terminal):
    print("\n================", terminal.goose_out_param[5], "==============")
    print(terminal.ied.get('desc', "ХЗ"))
    # for k, v in terminal.goose_out_param.items():
    #     print(f"GOOSEOutParam-{k}: {v}")
    for k, v in terminal.goose_out_data.items():
        # print(f"К набору данных ind{k}, привязан внутренний сигнал Экры {v}", terminal.signal_names.get(int(v), "В ЭКРАвских сидах не найти, бери мануал((("))
        print("----------------------------------------------------------")
        print(f"|ind{k}\t\t\t|{v}\t\t\t|", "Сигнал:", terminal.signal_names.get(int(v), "В ЭКРАвских сидах не найти, бери мануал((("))
        for subscriber in terminal.subscribers[k]:
            print("\t\t----------------------------------------------------------")
            print(f"|\t|\t| Подписчик: {subscriber[0]}, ind{subscriber[1]}")
    # for k, v in terminal.subscribers.items():
    #     print(f"На набор данных ind{k} подписаны {v}")
    if terminal.errors:
        print("!!!!!Неправильна привязка!!!!!")
    for err in terminal.errors:
        print(f"На несуществующий ind{err[0]} привязан {err[1]} ind{err[2]}")
    # for i in range(len(terminal.goose_in)):
    #     #print(terminal.goose_in[i][2])
    #     if terminal.goose_in[i][2] != "000000000000":
    #         for j in range(11):
    #             print(f"GOOSEIn-{i}-Param-{j}", terminal.goose_in[i][j])
    print("\n")

def print_terminal_in_file(terminal: Terminal, file: str):
    with open(file, 'a') as f:
        print("\n================", terminal.goose_out_param[5], "==============", file=f)
        print(terminal.ied.get('desc', "ХЗ"), file=f)
        # for k, v in terminal.goose_out_param.items():
        #     print(f"GOOSEOutParam-{k}: {v}")
        for k, v in terminal.goose_out_data.items():
            # print(f"К набору данных ind{k}, привязан внутренний сигнал Экры {v}", terminal.signal_names.get(int(v), "В ЭКРАвских сидах не найти, бери мануал((("))
            print("----------------------------------------------------------", file=f)
            print(f"|ind{k}\t|{v}\t|", "Сигнал:", terminal.signal_names.get(int(v), "В ЭКРАвских сидах не найти, бери мануал((("), file=f)
            for subscriber in terminal.subscribers[k]:
                print("\t\t----------------------------------------------------------", file=f)
                print(f"|\t\t|\t| Подписчик: {subscriber[0]}, ind{subscriber[1]}", file=f)
        # for k, v in terminal.subscribers.items():
        #     print(f"На набор данных ind{k} подписаны {v}")
        if terminal.errors:
            print("!!!!!Неправильна привязка!!!!!", file=f)
        for err in terminal.errors:
            print(f"На несуществующий ind{err[0]} привязан {err[1]} ind{err[2]}", file=f)
        # for i in range(len(terminal.goose_in)):
        #     #print(terminal.goose_in[i][2])
        #     if terminal.goose_in[i][2] != "000000000000":
        #         for j in range(11):
        #             print(f"GOOSEIn-{i}-Param-{j}", terminal.goose_in[i][j])
        print("\n", file=f)

def make_terminal_ekra(file_name: str):
    terminal = Terminal()
    tree = ET.parse(file_name)
    # print("\n")
    print("\n Читаем ",file_name)
    terminal.file_name = file_name
    root = tree.getroot()
    is_goose = False
    is_ekra = False
    for ied in root.findall(f"{IED}"):
        for k, v  in ied.attrib.items():
            print(k, v)
            terminal.ied[k] = v
    if terminal.ied["manufacturer"] != "EKRA":
        return None
    for p in root.findall(f"{COMMUNICATION}/{SUBNETWORK}/{CONNECTEDAP}/{ADDRESS}/{P}"):
        if "IP" == p.attrib["type"]:
            terminal.communication["IP"] = p.text
    for private in root.findall(f"{IED}/{PRIVATE}"): # пробегаю по параметрам исходящего GOOSE
        if "EKRA-GOOSEOutParam-" in private.attrib["type"]:
            is_goose = True
            #print(file_name, "{IED}/{PRIVATE}")
            if private.attrib["type"] == "EKRA-GOOSEOutParam-all": # В ПАДСах отличается
                #print(file_name, "__________", private.text)
                for p in private.text.split(";"):
                    p = p.split(":")
                    if len(p) == 2:
                        terminal.goose_out_param[int(p[0][1:])] = p[1]
                        #print(p)
                #print(terminal.goose_out_param)
            else:
                terminal.goose_out_param[int(private.attrib["type"][19:])] = private.text    
            #print(terminal.goose_out_param[int(private.attrib["type"][19:])])
        elif "EKRA-GOOSEIn-" in private.attrib["type"]  and "-Param-" in private.attrib["type"]:
            #print(2)
            tt = private.attrib["type"][13:].split("-Param-")
            tt[0], tt[1] = int(tt[0]), int(tt[1])
            # goose_in_x = []
            # for i in range(11):
            # print(tt)
            # print(terminal.goose_in.setdefault(tt[0], [None] * 11))
            terminal.goose_in.setdefault(tt[0], [None] * 11)[tt[1]] = private.text
            # terminal.goose_in[tt[0]][tt[1]] = private.text
    for private in root.findall(f"{IED}/{ACCESSPOINT}/{SERVER}/{LDEVICE}/{LN0}/{DATASET}/{PRIVATE}"):
        if "type" in private.attrib:
            if private.attrib["type"] == "EKRA-DSNum-all": # В ПАДСах отличается
                # print(file_name, "__________")
                for p in private.text.split(";"):
                    p = p.split(":")
                    if len(p) == 2:
                        #print(p)
                        terminal.goose_out_data[int(p[0][1:])] = p[1]
            else:
                terminal.goose_out_data[int(private.attrib["type"][11:])] = private.text
    for g in terminal.goose_out_data:
        terminal.subscribers[g] = []

    for dai in root.findall(f"{IED}/{ACCESSPOINT}/{SERVER}/{LDEVICE}/{LN}/{DOI}/{DAI}"):
        if "desc" in dai.attrib:
            try:
                k, v = dai.attrib["desc"].split(" - ", 1)
                #print(k, v)
                terminal.signal_names[int(k)] = v
            except:
                #print("Не смог записать в signal_names", dai.attrib["desc"])
                pass
    #print(terminal.goose_in)
    if is_goose:
        return terminal
    else:
        return None

def make_substation(cids: str):
    substation: Terminal = []
    with ZipFile(cids, mode='r') as zip_file:
        info = zip_file.namelist() 
        zip_file.extractall(cids.removesuffix(".zip"))
        for file in info:
            #print(file)
            #t = make_terminal(file) # сооздаю терминал
            if whose_cid(cids.removesuffix(".zip") + "/" + file) == "EKRA":
                ter = make_terminal_ekra(cids.removesuffix(".zip") + "/" + file)
                if ter:
                    substation.append(ter) # добовляю терминал в подстанцию
                #print_terminal(t) # распечатываю терминал для проверки. на этом этапе все верно

    # print(len(substation)) # убеждаюсь что на подстанции 2 терминала
    #[print_terminal(t) for t in substation] # распечатываю обатерминала. Почему то они одинаковые, как второй файл а архиве, оба как t  на второй итерации цикла

    for ter_out in substation: # бегу по терминалам подстанции
        # print(ter_out.file_name, ter_out.goose_out_param)
        # print(ter_out.file_name, "Исходящий гусь", ter_out.goose_out_param[4])
        print(ter_out.communication["IP"])
        appid_goose = ter_out.goose_out_param[4] # уникальный номер изходящего гуся
        quality = int(ter_out.goose_out_param[8]) # качество 0 - нет, 1 - до, 2 - после
        #quality = 0 # качество 0 - нет, 1 - до, 2 - после
        for ter_in in substation: # бегу по терминалам подстанции
            for goose_in_x in ter_in.goose_in.items(): # 
                if goose_in_x[1][3] == appid_goose: # 
                    if quality == 0:
                        index = int(goose_in_x[1][6])
                    elif quality == 1:
                        index = int(goose_in_x[1][6]) / 2
                    elif quality == 2:
                        index = (int(goose_in_x[1][6]) + 1) / 2
                    try:
                        ter_out.subscribers[index].append((ter_in.goose_out_param[5], int(goose_in_x[0]) + 1))
                    except:
                        ter_out.errors.append((index, ter_in.goose_out_param[5], int(goose_in_x[0]) + 1))

    substation.sort(key=lambda x: x.communication["IP"])                
    # [print_terminal(t) for t in substation]
    [print_terminal_in_file(t, cids.rstrip(".zip") + ".txt") for t in substation] 
    make_xl(substation, cids)
    empty_vertical, empty_horizontal = fill_xl(cids.removesuffix(".zip") + ".xlsx", substation)
    #paint_xl(cids.removesuffix(".zip") + ".xlsx")
    paint_xl(cids.removesuffix(".zip") + ".xlsx", empty_vertical, empty_horizontal)

def whose_cid(cid: str):
    print("\n!!!", cid)
    try:
        tree = ET.parse(cid)
    except:
        p = cid.rsplit("/", 2)
        print("Не удалось открыть", cid)
        err = open (p[0] + '/' + "no_read.txt", 'a')
        err.write(p[-1])
        err.write("\n")
        err.close()
        return "Не удалось открыть"
    print("Определяем чей cid ",cid)
    root = tree.getroot()
    for header in root.findall(f"{HEADER}"):        
        if ("id", "")  in header.attrib.items():
            for ied_2 in root.findall(f"{IED}"):
                if ("manufacturer", "EKRA")  in ied_2.attrib.items():
                    print("This is EKRA")
                    return "EKRA"
        else:
            for ied_3 in root.findall(f"{IED}"):
                if ied_3.attrib["name"] == header.attrib["id"] and ied_3.attrib["manufacturer"] == "SIEMENS":
                    print("This is SIEMENS")
                    return "SIEMENS"
    return "Неизвестное устройство"


def make_matrix(substation):
    pass

HORIZ = 3
VERT = 3

def make_xl(substation, path):
    #print(substation[0].signal_names)
    wb = openpyxl.Workbook()
    ws = wb.active
    # HORIZ = 3
    # VERT = 2
    horiz = HORIZ + 1 # первый столбец для заполнения горизонтали терминалов
    vert = VERT + 1 # первая строка для заполнения вертикали терминалов
    for t in substation:
        print("заполняем  xl", t.goose_out_param[5])
        ws.cell(VERT, horiz).value = t.goose_out_param[5]
        ws.cell(VERT, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
        ws.cell(VERT - 2, horiz).value = t.communication["IP"] 
        ws.cell(VERT - 2, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)

        letter = ws.cell(VERT, horiz).column_letter # получаем букву текущей ячейки чтобы изменить ее ширину
        ws.column_dimensions[letter].width = 3 # меняем ширину текущего столбца
        horiz += 1
        start_group = ws.cell(VERT, horiz).column_letter # получаем букву текущей для группировки ячеек
        end_group = None # если у терминала не будет входящих гусей, то конец гусей не определиться и колонки не будут группироваться
        #print(t.goose_in)
        for ind, goose in t.goose_in.items():
            #print(ind, goose)
            if goose[3] == '0':
                continue
            ws.cell(VERT - 1, horiz).value = goose[0]
            ws.cell(VERT, horiz).value = "ind_" + str(ind + 1)
            ws.cell(VERT, horiz).alignment = openpyxl.styles.Alignment(textRotation=90)
            letter = ws.cell(VERT, horiz).column_letter # получаем букву текущей ячейки чтобы изменить ее ширину
            ws.column_dimensions[letter].width = 3 # меняем ширину текущего столбца
            end_group = ws.cell(VERT, horiz).column_letter
            horiz += 1
        if end_group:
            ws.column_dimensions.group(start_group, end_group, hidden=True)
        ws.cell(vert, HORIZ).value = t.goose_out_param[5]
        ws.cell(vert, HORIZ - 2).value = t.communication["IP"] 
        vert += 1
        start_group = vert
        end_group = None
        for ind, signal in t.goose_out_data.items():
            ws.cell(vert, HORIZ).value = "ind_" + str(ind)
            ws.cell(vert, HORIZ - 1).value = signal
            ws.cell(vert, HORIZ - 2).value = t.signal_names.get(int(signal), "В ЭКРАвских сидах не найти, бери мануал(((")
            end_group = vert
            vert += 1
        if end_group:
            ws.row_dimensions.group(start_group, end_group, hidden=True)
    

    ws.column_dimensions.group('A', ws.cell(VERT, HORIZ - 1).column_letter, hidden=True)
    ws.row_dimensions.group(1, VERT - 1, hidden=True)
    ws.column_dimensions[ws.cell(VERT, HORIZ - 2).column_letter].width = 22
    ws.column_dimensions[ws.cell(VERT, HORIZ - 1).column_letter].width = 5
    ws.column_dimensions[ws.cell(VERT, HORIZ).column_letter].width = 20
    
    
    ws.row_dimensions[VERT].height = 100
    ws.row_dimensions[VERT - 2].height = 110
    ws.freeze_panes = ws.cell(VERT + 1, HORIZ + 1)
    wb.save(path.removesuffix(".zip") + ".xlsx")

def fill_xl(file_xl, substation):
    wb = openpyxl.load_workbook(file_xl)
    ws = wb.active
    # HORIZ = 3
    # VERT = 2
    empty_vertical = [True] * (ws.max_row + 1)
    empty_horizontal = [True] * (ws.max_column + 1)
    horiz = HORIZ + 1 # первый столбец для заполнения пересечений
    vert = VERT + 1# первая строка для заполнения пересечений
    for t in substation:
        print("пересечения в xl", t.goose_out_param[5])
        # этим циклом получаю множество подписчиков терминала
        subscribers_term = set()
        for a in t.subscribers.values():
            if a:
                for b in a:
                    subscribers_term.add(b[0])
        # заполняем пересечения самих терминалов
        for y in range(horiz, ws.max_column + 1):
            if ws.cell(VERT, y).value in subscribers_term:
                ws.cell(vert, y).value = "X"
                empty_vertical[vert] = False
                empty_horizontal[y] = False
        vert += 1
        # заполняем пересечения ind_x
        for ind_out, subscribers in t.subscribers.items():
            #print(subscribers)
            for y in range(horiz, ws.max_column + 1):
                if ws.cell(VERT, y).value.startswith("ind_"):
                    ind_in = int(ws.cell(VERT, y).value.removeprefix("ind_"))
                    for subscriber in subscribers:
                        if subscriber[0] == current_terminal and subscriber[1] == ind_in:
                            if ws.cell(VERT - 1, y).value == "1": # если входящий гусь включен
                                ws.cell(vert, y).value = "Да"
                                empty_vertical[vert] = False
                                empty_horizontal[y] = False
                            elif ws.cell(VERT - 1, y).value == "0": # если входящий гусь отключен
                                ws.cell(vert, y).value = "Нет"
                                empty_vertical[vert] = False
                                empty_horizontal[y] = False
                else:
                    current_terminal = ws.cell(VERT, y).value
            #print(int(ws.cell(vert, HORIZ).value.removeprefix("ind_")))
            vert += 1
    wb.save(file_xl)
    return empty_vertical, empty_horizontal

def paint_xl(file_xl, empty_vertical, empty_horizontal):
    
    wb = openpyxl.load_workbook(file_xl)
    ws = wb.active
    # HORIZ = 3
    # VERT = 2
    horiz = HORIZ + 1 # первый столбец для заполнения пересечений
    vert = VERT + 1# первая строка для заполнения пересечений

    print("красим входящие, которые не подписаны")
    for x in range(horiz, ws.max_column + 1):        
        # is_empty_column = True
        # for y in range(vert, ws.max_row + 1):
        #     if ws.cell(y, x).value != None:
        #         is_empty_column = False
        #         break
        if empty_horizontal[x]:
            # for rows in ws.iter_rows(min_row=vert, max_row=ws.max_row, min_col=x, max_col=x):
            #     for cell in rows:
            #         cell.fill = PatternFill('solid', fgColor="ffffff")
            ws.cell(VERT, x).fill = PatternFill('solid', fgColor="8A3324")
    print("красим исходящие, на которые подписаны")
    for y in range(vert, ws.max_row + 1):        
        # is_empty_column = True
        # for x in range(horiz, ws.max_column + 1):
        #     if ws.cell(y, x).value != None:
        #         is_empty_column = False
        #         break
        if empty_vertical[y]:
            # for rows in ws.iter_rows(min_row=y, max_row=y, min_col=horiz, max_col=ws.max_column):
            #     for cell in rows:
            #         cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type = "solid")
            ws.cell(y, HORIZ).fill = PatternFill('solid', fgColor="8A3324")
    wb.save(file_xl)


# def paint_xl(file_xl):
    
#     wb = openpyxl.load_workbook(file_xl)
#     ws = wb.active
#     # HORIZ = 3
#     # VERT = 2
#     horiz = HORIZ + 1 # первый столбец для заполнения пересечений
#     vert = VERT + 1# первая строка для заполнения пересечений

#     print("красим входящие, которые не подписаны")
#     for x in range(horiz, ws.max_column + 1):        
#         is_empty_column = True
#         for y in range(vert, ws.max_row + 1):
#             if ws.cell(y, x).value != None:
#                 is_empty_column = False
#                 break
#         if is_empty_column:
#             # for rows in ws.iter_rows(min_row=vert, max_row=ws.max_row, min_col=x, max_col=x):
#             #     for cell in rows:
#             #         cell.fill = PatternFill('solid', fgColor="ffffff")
#             ws.cell(VERT, x).fill = PatternFill('solid', fgColor="8A3324")
#     print("красим исходящие, на которые подписаны")
#     for y in range(vert, ws.max_row + 1):        
#         is_empty_column = True
#         for x in range(horiz, ws.max_column + 1):
#             if ws.cell(y, x).value != None:
#                 is_empty_column = False
#                 break
#         if is_empty_column:
#             # for rows in ws.iter_rows(min_row=y, max_row=y, min_col=horiz, max_col=ws.max_column):
#             #     for cell in rows:
#             #         cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type = "solid")
#             ws.cell(y, HORIZ).fill = PatternFill('solid', fgColor="8A3324")
#     wb.save(file_xl)

class SuperSocket():
    def __init__(self, sock):
        self._sock = sock

    def send_msg(self, msg):
        # Каждое сообщение будет иметь префикс в 4 байта блинной(network byte order)
        msg = struct.pack('>I', len(msg)) + msg
        self._sock.send(msg)

    def recv_msg(self):
        # Получение длины сообщения и распаковка в integer
        raw_msglen = self.recvall(4)
        if not raw_msglen:
            return None
        msglen = struct.unpack('>I', raw_msglen)[0]
        # Получение данных
        return self.recvall(msglen)

    def recvall(self, n):
        # Функция для получения n байт или возврата None если получен EOF
        data = b''
        while len(data) < n:
            packet = self._sock.recv(n - len(data))
            if not packet:
                return None
            data += packet
        return data

sock = socket.socket()
# IP = "213.178.155.215"
IP = ""
PORT = 51085
sock.bind((IP, PORT))
sock.listen(5)

while True:
    print("start")
    # начинаем принимать соединения
    conn, addr = sock.accept()
    # выводим информацию о подключении
    print('connected:', addr)
    super_sock = SuperSocket(conn)
     # получаем название файла
    name_f = super_sock.recv_msg().decode('UTF-8')
    print("Имя архива (str)", name_f)
    path = "clients/" + name_f.removesuffix(".zip") + "/"
    name = name_f.removesuffix(".zip")
    # создаем каталог для файлов клиента
    # print(type("clients/" + name_f.strip(".zip")))
    # print("clients/" + name_f.strip(".zip"))
    if not os.path.isdir("clients/" + name_f.removesuffix(".zip")):
        os.mkdir("clients/" + name_f.removesuffix(".zip"))
    # открываем файл в режиме байтовой записи в папке клиента
    f = open("clients/" + name_f.removesuffix(".zip") + "/" + name_f, 'wb')
    f.write(super_sock.recv_msg())
    f.close()

    print("Обрабытываем сиды")
    # make_substation("clients/" + name_f.strip(".zip") + "/" + name_f)
    make_substation(path + name_f)
    
    # name_f = "clients/" + name_f.strip(".zip") + "/" + name_f.strip(".zip") + ".txt"
    name_f = path + name + ".xlsx"
    print("Отправляем", name_f)
    super_sock.send_msg((name + ".xlsx").encode('UTF-8'))
    f = open (name_f, "rb")
    super_sock.send_msg(f.read())
    f.close()

    conn.close()
    print("end")
    print("убери break после отладки")
    break

